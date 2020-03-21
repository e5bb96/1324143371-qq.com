from openpyxl import load_workbook # 导入读取excel文件的模块
class append_:
    def __init__(self):
        names = input('请输入要录入的文件名称：')
        read = load_workbook(r'E:\wehath\WeChat Files\zx1324143371\FileStorage\File\2019-10\{}.xlsx'.format(names))
        active_sheet = read.active
        self.lists = []
        for i in range(1,11):
            hang = active_sheet.cell(row=2,column=i).value
            if i == 7:
                print(hang)        
            self.lists.append(hang)
        print(self.lists)

    def xlsx_(self):
        rows = int(input('请输入要添加到xlsx的文件行数（如1,2,3,4....）'))
        read = load_workbook(r'E:\wehath\WeChat Files\zx1324143371\FileStorage\File\2019-10\基础信息表(例表).xlsx')
        appends = read.active
        for i in range(1,11):
            n = i - 1
            appends.cell(row = rows,column = i,value=self.lists[n])

        read.save(r'E:\wehath\WeChat Files\zx1324143371\FileStorage\File\2019-10\基础信息表(例表).xlsx')
print('录入完成')
opens = append_()
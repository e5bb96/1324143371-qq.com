  # r代表read阅读，也可w重写入，a追加
  # r代表read阅读，也可w重写入，a追加
  # r代表read阅读，也可w重写入，a追加
file1 = open (r'C:\python\p.py\123.txt','w',encoding='gbk')
f = file1.write('我爱你\n')
a = file1.write('机油cao')
file1.close()
file2 = open(r'C:\python\p.py\123.txt','r',encoding='gbk')
c = file2.read()
print(c)
file2.close()


# read        读取整个文件
# readline    读取下一行
# readlines   读取整个文件到一个迭代器以供我们遍历（读取到一个list中，以供使用，比较方便）
with open(r'C:\python\p.py\123.txt','a') as work:
    works = work.write('gbk')
with open(r'C:\python\p.py\123.txt','r') as w:
    ws = w.readlines()
    print(ws)
    for i in ws:
        print(i)

# Console.WriteLine()方法是将要输出的字符串与换行控制字符一起输出,当次语句执行完毕时,光标会移到目前输出字符串的下一行.
# 至于Console.Write()方法,光标会停在输出字符串的最后一个字符后,不会移动到下一行。
# file.write(str)的参数是一个字符串，就是你要写入文件的内容.
# file.writelines(sequence)的参数是序列，比如列表，它会迭代帮你写入文件。

ping = []
with open (r'C:\python\p.py\1234.txt','w',encoding='gbk') as file:
    files = file.write('罗恩 23 35 44\n哈利 60 77 68 88 90\n赫敏 97 99 89 91 95 99\n马尔福 100 85 90')
with open(r'C:\python\p.py\1234.txt','r') as file :
    files = file.readlines()
    print(files)

    for i in files:
        date = i.split()  #把字符串切分为更细的一个个字符
        for k in date[1:]:
            sum = int(k)
        result = date[0] + str(sum) + '\n'
        print(result)
        ping.append(result)
print(ping)
    # for i in files:   # str.join(列表) 将列表中字符串间增加(str)字符
    #     date = i.split()
    #     dates = date[1:]
    #     b = '+'
    #     datess = b.join(dates)
    #     print(datess)

with open (r'C:/python/p.py/1234.txt','w') as winner:
    winners = winner.writelines(ping)
with open(r'C:\python\p.py\1234.txt','r') as file :
    files = file.readlines()
    print(files)

if __name__ == "__main__":
    pass
# class num:
#     biger = '9'
#     def number(self):
#         print(1314)
# def nums():
#     small = '1'
#     return small
# coss = 4

#①
# import story
# n=story.num()
# print(n.biger)
# print(story.nums())
# print(story.coss)

#②
# import story as s  #  将story命名为s
# print(s.coss)

#③  from ... import ...
# from story import num,nums
# a = num()
# a.number()
# print(a.biger)
# print(nums())




#csv
import csv
with open (r'C:\Users\e5bb96\Desktop\names.csv','r',encoding='utf-8') as f:
    writer = csv.reader(f)
    for i in writer:
        print(i)

#'a'追加
with open (r'C:\Users\e5bb96\Desktop\names.csv','a',encoding='utf-8') as f:
    a = csv.writer(f)
    a.writerow(['4','e5bb96','886'])



file=open(r'C:\Users\e5bb96\Desktop\names.csv','a')
file.write('2312,adsad,2321fcd')        #逗号','相当于excel中每个空格的隔离符号
file.close









import requests
from bs4 import BeautifulSoup
res = requests.get('https://localprod.pandateacher.com/python-manuscript/crawler-html/spider-men5.0.html')
print(res.status_code)  #打印变量态度码，已检查请求是否成功
# pic = res.content   #把Reponse对象内容（图片，音乐）以二进阶数据的形式返回
# photo = open(r'D:\pyhton\ppt.jpg','wb')
# photo.write(pic)
# photo.close()


res.encoding='utd-8'
novel = res.text   #把Response对象以字符串的形式返回
#print(novel[:100])
soup = BeautifulSoup(novel,'html.parser')  #把网页解析为BeautifulSoup对象
items = soup.find_all(class_='books')  
for item in items:
    kind = item.find('h2')
    title = item.find(class_='title')
    print(kind,'\n',title)





from openpyxl import Workbook # 导入新建excel文件的模块
from openpyxl import load_workbook # 导入读取excel文件的模块


xls_read = load_workbook(r'E:\new\list.xlsx')  # 打开excel文件名为'****.xlsx'
print(xls_read.sheetnames)  # 查看工作表'.xlsx'中的所有sheet名，以列表形式生成
print(xls_read.active)    # 查看文件pyxl_test的活动中sheet
xls_read.active.title = 'test'  # 将活动中的sheet名称变更为test
xls_read_sheet = xls_read.active    # 将活动中的sheet赋值给变量
xls_read_sheet = xls_read.get_sheet_by_name('test')     # 获取excel文件的某一个sheet
sheet1_max_colum = xls_read_sheet.max_column  # 获取最大行数 结果：10
a = xls_read['text']    #选择其中一个工作表


row=[['美国队长','刚喋血','大数据线'],['郑晓纯','大闸蟹','额滴神啊']]
for i in row:   #多行输入
    xls_read_sheet.append(row)       #在表格中添加一行数据

xls_read_sheet['A1']='漫威'     #给'A1'单元格写入数据
print(xls_read_sheet['C'])    # 读取sheet中的C列所有数据，数据是以元组形式呈现
print(xls_read_sheet['A1'].value) # 读取sheet中'A1'单元格的值
print(xls_read_sheet.max_column)  # 查看sheet中最大的列合计值，统计依据是只要单元格含有值，就算一列
print(xls_read_sheet.max_row)     # 查看sheet中最大的行合计值，统计依据是只要单元格含有值，就算一行
